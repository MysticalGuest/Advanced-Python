import openpyxl

# wb = openpyxl.Workbook()
# wb.create_sheet('learn_openpyxl')
# wb.save('courses.xlsx')

# 读取excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook('courses.xlsx')
# 第二步：选取表单
sh = wb['Sheet']
# sh = wb['learn_openpyxl']
# 第三步：读取数据
# 参数 row:行  column:列
ce = sh.cell(row=1, column=1)   # 读取第一行，第一列的数据
print("读取第一行，第一列的数据: "+ce.value)
# 按行读取数据 list(sh.rows)
print("按行读取数据: ")
print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据

for cases in list(sh.rows)[1:]:
    # data1 = cases[0].value
    # data2 = cases[1].value
    # data3 = cases[2].value
    # data4 = cases[3].value
    # print(data1, data2, data3, data4)
    for data in cases:
        print(data.value, end="  ")
    print()
# 关闭工作薄
wb.close()
